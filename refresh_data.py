#!/usr/bin/env python3
"""
Duke IM Resident Dashboard — Weekly Data Refresh
=================================================
Reads the team points Excel workbook and regenerates data.js for the dashboard.

SETUP (one-time)
----------------
1. Install dependency:
       pip3 install openpyxl

2. Set EXCEL_FILE below to the path of your workbook.
   For a file synced via OneDrive / Box / Dropbox, use the local sync path —
   the file itself never leaves your machine.

3. Run it manually whenever the spreadsheet is updated:
       python3 refresh_data.py

4. Commit the updated data.js to push live changes to GitHub Pages.

EXCEL WORKBOOK FORMAT
---------------------
Sheet tab named "Points" (or first sheet), headers in row 1:

    Date | Resident | Team | Category | Points | Notes

    Date:     Any Excel date cell, or text as YYYY-MM-DD or M/D/YYYY
    Resident: Optional. Full name matching a member in teams.html's ROSTER —
              if given and Team is left blank, the team is looked up
              automatically. Use this instead of an in-sheet lookup formula
              (e.g. XLOOKUP): openpyxl never evaluates formulas, and it also
              strips any cached formula value on save (write_attendance_summary
              below resaves this workbook every run), so formula-based Team
              cells silently go blank.
    Team:     Must match a key in TEAM_COLORS below (e.g. "Gold", "Blue").
              Required only if Resident is blank (e.g. team-wide bonus rows).
    Category: Must match a key in CATEGORY_COLORS below
    Points:   A number
    Notes:    Optional — ignored by the dashboard

An optional "Attendance" sheet (written by scrape_attendance.py) is also
folded in automatically if present — see ATTENDANCE_SHEET_NAME/EVENT_POINTS
below. Run scrape_attendance.py first to sync it, then run this script.
"""

import json
import math
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit('openpyxl not found. Install it with:  pip3 install openpyxl')

# ============================================================
# CONFIGURATION — edit these to match your setup
# ============================================================

# Full path to the shared Excel workbook.
# Use the local OneDrive / Box / Dropbox sync path if the file is shared that way.
EXCEL_FILE = Path('/Users/nbrazeau/Library/CloudStorage/OneDrive-SharedLibraries-DukeUniversity/Duke Chiefs 2024-2026 - Documents/zChief_Gamification/Point_Spreadsheet.xlsx')


# Sheet tab name with the points log. None = use the first (active) sheet —
# only safe with a single-sheet workbook. Now that the workbook has multiple
# tabs (OtherPoints, Residents, Teams, Categories, AttendancePoints), hardcode
# this so it doesn't depend on whichever tab Excel last had "active" when saved.
SHEET_NAME = 'OtherPoints'

# Sheet tab name with the scraped attendance roster (written by scrape_attendance.py).
# Skipped entirely if this sheet doesn't exist yet.
ATTENDANCE_SHEET_NAME = 'AttendancePoints'

# Sheet tab (re)written each run with one row per resident's cumulative
# attendance points, sourced from ATTENDANCE_SHEET_NAME.
ATTENDANCE_SUMMARY_SHEET_NAME = 'Attendance Summary'

# Points awarded per attendance event type. Add new event types here as needed.
EVENT_POINTS = {
    'Noon Conference':  20,
    'Learning Session': 10,
    'Welcome':          20,
}

# teams.html holds the single source of truth for resident team membership
# (the ROSTER array) — parsed at runtime instead of duplicated here.
TEAMS_HTML = Path(__file__).parent / 'teams.html'

# Team colors — keys must match the Team column values exactly (case-sensitive).
TEAM_COLORS = {
    'Creatininjas':    '#2a78d6',
    'Scopetrotters':   '#1baf7a',
    'PEEPs':           '#eda100',
    'Karius':          '#008300',
    'Hemoglobbers':    '#4a3aa7',
    'Stentinels':      '#e34948',
    'Jointventurers':  '#e87ba4',
    'Codeblazers':     '#eb6834',
    'Glandiators':     '#0891b2',
    'Remissionaries':  '#9d174d',
}

# Category colors — keys must match Category column values exactly.
# 'All Points' is used automatically when no Category column exists.
CATEGORY_COLORS = {
    'Mentoring':      '#012169',
    'Teaching':       '#C84E00',
    'Diagnosing':     '#339898',
    'Grittiness':     '#D97706',
    'Follow-Through': '#DC2626',
    'Wellness':       '#16A34A',
    'Learning':       '#6366F1',
    'Attendance':     '#EC4899',
    'All Points':     '#012169',
}

# Academic year starts in July.
ACADEMIC_MONTHS = ['Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun']

# Academic year start date — used to bucket events into weeks 1-52 (matches the
# "Week X of 52" calculation in index.html).
ACADEMIC_YEAR_START = datetime(2026, 7, 1)

# Output path — always the data.js sitting next to this script.
DATA_JS = Path(__file__).parent / 'data.js'

# ============================================================


def parse_date(value):
    if isinstance(value, datetime):
        return value
    if hasattr(value, 'year'):
        return datetime(value.year, value.month, value.day)
    s = str(value).strip()
    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y', '%-m/%-d/%Y', '%-m/%-d/%y'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    raise ValueError(f'Cannot parse date: {value!r}')


def normalize_name(name):
    """Lowercase and collapse whitespace for roster matching."""
    return re.sub(r'\s+', ' ', str(name or '')).strip().lower()


def build_name_index(roster_map):
    """normalized name -> (canonical roster name, team).

    Names in the spreadsheet are hand-typed, so 'maxwell sumner' and
    'Sean Lafata' still have to match the roster's 'Maxwell Sumner' /
    'Sean LaFata'.
    """
    return {normalize_name(n): (n, t) for n, t in roster_map.items()}


def read_events(path, sheet_name, roster_map=None):
    """Read point events from the OtherPoints sheet.

    The Team column is optional per-row: if it's blank but a Resident name is
    given, the team is looked up from roster_map (parsed from teams.html) —
    this avoids relying on an in-sheet formula (e.g. XLOOKUP), whose cached
    value openpyxl can't compute and will blank out on any resave.
    """
    roster_map = roster_map or {}
    name_index = build_name_index(roster_map)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return []

    headers = [str(h).strip().lower() if h is not None else '' for h in rows[0]]

    def col(name, required=True):
        try:
            return headers.index(name)
        except ValueError:
            if required:
                raise ValueError(f'Column "{name}" not found. Headers detected: {list(rows[0])}')
            return None

    ci = {
        'date':     col('date'),
        'team':     col('team', required=False),      # optional if 'resident' is given
        'resident': col('resident', required=False),  # optional — used to derive team
        'category': col('category', required=False),  # optional
        'points':   col('points'),
    }

    events = []
    for row in rows[1:]:
        try:
            date     = parse_date(row[ci['date']])
            team     = str(row[ci['team']] or '').strip() if ci['team'] is not None else ''
            resident = str(row[ci['resident']] or '').strip() if ci['resident'] is not None else ''
            cat      = str(row[ci['category']] or '').strip() if ci['category'] is not None else 'All Points'
            pts      = float(row[ci['points']] or 0)
        except (TypeError, ValueError, IndexError):
            continue

        if not team and resident:
            _, team = name_index.get(normalize_name(resident), (None, ''))
            if not team and pts:
                print(f'  [points] skipping unmapped resident: {resident!r}')

        if team and pts:
            events.append({'date': date, 'team': team, 'category': cat, 'points': pts})

    return events


def load_roster_from_teams_html(path):
    """Parse the ROSTER array out of teams.html and return {full_name: team_name}."""
    text = path.read_text(encoding='utf-8')
    roster = {}
    for team_match in re.finditer(r"team:\s*'([^']+)'.*?members:\s*\[([^\]]*)\]", text, re.DOTALL):
        team_name = team_match.group(1)
        members_blob = team_match.group(2)
        for name_match in re.finditer(r"'([^']+)'", members_blob):
            roster[name_match.group(1)] = team_name
    return roster


def read_attendance(path, sheet_name, roster_map):
    """Read the scraped Attendance sheet.

    Returns (events, resident_totals):
      - events: list of {'date', 'team', 'category': 'Attendance', 'points'} for aggregate()
      - resident_totals: {name: cumulative_attendance_points} for every resident in
        roster_map (0 for residents with no attendance rows yet)
    """
    resident_totals = {name: 0 for name in roster_map}
    name_index = build_name_index(roster_map)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        wb.close()
        return [], resident_totals

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return [], resident_totals

    headers = [str(h).strip().lower() if h is not None else '' for h in rows[0]]

    def col(name):
        try:
            return headers.index(name)
        except ValueError:
            raise ValueError(f'Column "{name}" not found in Attendance sheet. Headers detected: {list(rows[0])}')

    date_i = col('date')
    name_i = col('name')
    event_i = col('event')

    events = []
    for row in rows[1:]:
        try:
            date = parse_date(row[date_i])
            name = str(row[name_i] or '').strip()
            event = str(row[event_i] or '').strip()
        except (TypeError, ValueError, IndexError):
            continue

        canonical, team = name_index.get(normalize_name(name), (None, None))
        if not team:
            print(f'  [attendance] skipping unmapped name: {name!r}')
            continue

        points = EVENT_POINTS.get(event)
        if not points:
            print(f'  [attendance] skipping unknown event type: {event!r}')
            continue

        events.append({'date': date, 'team': team, 'category': 'Attendance', 'points': points})
        resident_totals[canonical] = resident_totals.get(canonical, 0) + points

    return events, resident_totals


def write_attendance_summary(path, summary_sheet_name, resident_totals, roster_map):
    """(Re)write the per-resident cumulative attendance points sheet."""
    wb = openpyxl.load_workbook(path)

    if summary_sheet_name in wb.sheetnames:
        del wb[summary_sheet_name]
    ws = wb.create_sheet(summary_sheet_name, 0)

    ws.append(['Name', 'Team', 'Attendance Points'])
    for name in sorted(roster_map, key=lambda n: (roster_map[n], n)):
        ws.append([name, roster_map[name], resident_totals.get(name, 0)])

    wb.save(path)
    return len(roster_map)


def week_number(date):
    """1-52 week index since ACADEMIC_YEAR_START, matching the frontend's calc."""
    delta_days = (date - ACADEMIC_YEAR_START).days + 1
    return min(52, max(1, math.ceil(delta_days / 7)))


def week_month_label(week_num):
    """Calendar month abbreviation that a given academic week falls in."""
    week_start = ACADEMIC_YEAR_START + timedelta(weeks=week_num - 1)
    return week_start.strftime('%b')


def aggregate(events, today=None):
    """Roll events up into the shape data.js expects.

    Each team also gets `rank` (current standing) and `prevRank` (standing as of
    the end of yesterday, computed from events dated strictly before today) so
    the dashboard can show a movement arrow. `prevRank` is None when there are
    no events before today at all — i.e. the season just started and there is no
    prior standing to compare against.
    """
    today = (today or datetime.now()).replace(hour=0, minute=0, second=0, microsecond=0)

    team_monthly = {}
    team_weekly  = {}
    cat_totals   = {}
    seen_months  = set()
    seen_weeks   = set()

    for e in events:
        month = e['date'].strftime('%b')
        seen_months.add(month)

        wk = week_number(e['date'])
        seen_weeks.add(wk)

        t = e['team']
        team_monthly.setdefault(t, {})
        team_monthly[t][month] = team_monthly[t].get(month, 0) + e['points']

        team_weekly.setdefault(t, {})
        team_weekly[t][wk] = team_weekly[t].get(wk, 0) + e['points']

        c = e['category']
        cat_totals[c] = cat_totals.get(c, 0) + e['points']

    months = [m for m in ACADEMIC_MONTHS if m in seen_months]
    weeks  = sorted(seen_weeks)
    week_months = [week_month_label(w) for w in weeks]

    teams = []
    for name in set(team_monthly) | set(team_weekly):
        monthly = [int(team_monthly.get(name, {}).get(m, 0)) for m in months]
        weekly  = [int(team_weekly.get(name, {}).get(w, 0)) for w in weeks]
        teams.append({
            'name':    'Team ' + name,
            'color':   TEAM_COLORS.get(name, '#6B7280'),
            'total':   sum(monthly),
            'monthly': monthly,
            'weekly':  weekly,
        })
    # Tie-break by name so equal totals rank in a stable order run-to-run —
    # otherwise tied teams could swap places on a rerun and fake a rank change.
    teams.sort(key=lambda t: (-t['total'], t['name']))
    for i, t in enumerate(teams):
        t['rank'] = i + 1

    prior_totals = {}
    for e in events:
        if e['date'] < today:
            key = 'Team ' + e['team']
            prior_totals[key] = prior_totals.get(key, 0) + e['points']

    if prior_totals:
        prior_order = sorted(
            (t['name'] for t in teams),
            key=lambda n: (-prior_totals.get(n, 0), n),
        )
        prior_rank = {name: i + 1 for i, name in enumerate(prior_order)}
        for t in teams:
            t['prevRank'] = prior_rank[t['name']]
    else:
        for t in teams:
            t['prevRank'] = None

    categories = [
        {
            'label': label,
            'value': int(v),
            'color': CATEGORY_COLORS.get(label, '#6B7280'),
        }
        for label, v in cat_totals.items()
    ]

    return {
        'months': months,
        'weeks': weeks,
        'weekMonths': week_months,
        'teams': teams,
        'categories': categories,
    }


def write_data_js(data):
    ts   = datetime.now().strftime('%Y-%m-%d %H:%M')
    blob = json.dumps(data, indent=4)
    content = (
        f'// Auto-generated {ts} by refresh_data.py — do not edit manually.\n'
        f'// To refresh: run python3 refresh_data.py then commit data.js\n'
        f'const SAMPLE_DATA = {blob};\n\n'
        f'async function loadDashboardData() {{\n'
        f'    return SAMPLE_DATA;\n'
        f'}}\n'
    )
    DATA_JS.write_text(content, encoding='utf-8')

    n_teams  = len(data['teams'])
    n_months = len(data['months'])
    n_pts    = sum(t['total'] for t in data['teams'])
    print(f'[{ts}] data.js updated — {n_teams} teams, {n_months} months, {n_pts:,} total pts.')


def diagnose(path, sheet_name):
    """Print sheet names, row count, headers, and first 5 data rows to help debug."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print(f'\n--- DIAGNOSTIC ---')
    print(f'Sheet tabs found:  {wb.sheetnames}')
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    print(f'Reading sheet:     "{ws.title}"')
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    print(f'Total rows read:   {len(rows)}')
    for i, row in enumerate(rows[:6], start=1):
        label = 'headers' if i == 1 else 'data   '
        print(f'  Row {i} ({label}): {list(row)}')
    print(f'------------------\n')


if __name__ == '__main__':
    if not EXCEL_FILE.exists():
        sys.exit(
            f'\nERROR: Excel file not found:\n  {EXCEL_FILE}\n\n'
            f'Edit EXCEL_FILE in refresh_data.py to point to your workbook.\n'
        )

    roster_map = load_roster_from_teams_html(TEAMS_HTML) if TEAMS_HTML.exists() else {}

    try:
        events = read_events(EXCEL_FILE, SHEET_NAME, roster_map)
    except ValueError as e:
        diagnose(EXCEL_FILE, SHEET_NAME)
        sys.exit(f'\nERROR reading workbook: {e}\n')
    except Exception as e:
        diagnose(EXCEL_FILE, SHEET_NAME)
        sys.exit(f'\nERROR: {e}\n')

    if not events:
        diagnose(EXCEL_FILE, SHEET_NAME)
        sys.exit(
            'ERROR: No valid data rows found.\n'
            'Check the diagnostic output above — look for:\n'
            '  - Required headers: Date, Points  (Category optional; either Team or a\n'
            '    Resident name matching teams.html is required)\n'
            '  - Date values parseable as dates (e.g. 7/1/2024 or 2024-07-01)\n'
            '  - Points column containing numbers\n'
            '  - Team values matching TEAM_COLORS keys in refresh_data.py\n'
        )

    if TEAMS_HTML.exists():
        attendance_events, resident_totals = read_attendance(EXCEL_FILE, ATTENDANCE_SHEET_NAME, roster_map)
        if attendance_events:
            print(f'[attendance] merged {len(attendance_events)} attendance event(s).')
        events += attendance_events

        n_residents = write_attendance_summary(EXCEL_FILE, ATTENDANCE_SUMMARY_SHEET_NAME, resident_totals, roster_map)
        print(f'[attendance] wrote "{ATTENDANCE_SUMMARY_SHEET_NAME}" sheet — {n_residents} resident row(s).')

    data = aggregate(events)
    write_data_js(data)
