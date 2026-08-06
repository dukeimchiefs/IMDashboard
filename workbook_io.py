#!/usr/bin/env python3
"""
Duke IM Resident Dashboard — Guarded Workbook Writes
====================================================
Both scrape_attendance.py and refresh_data.py save the shared
Point_Spreadsheet.xlsx, which lives on a OneDrive shared-library sync path.
A plain openpyxl `wb.save(path)` there is riskier than it looks:

  - openpyxl truncates and rewrites the file in place, so a crash or a
    interrupted write leaves a corrupt workbook with no way back.
  - If someone has the workbook open in Excel, or OneDrive is mid-sync, the
    cloud copy can win the resulting conflict and silently revert the save.
    That is how the August attendance rows went missing.
  - Every resave strips the cached values of in-sheet formulas (the XLOOKUP
    in the OtherPoints Team column), so needless writes actively degrade the
    workbook.

Every save goes through save_workbook() below, which skips writes that would
change nothing, refuses to write while Excel holds the file or when the file
changed underneath us, keeps a rolling local backup, swaps the new file in
atomically, and restores the backup if a sheet lost rows.

Backups live outside the synced folder and never leave the machine.
"""

import os
import shutil
import tempfile
from datetime import datetime
from pathlib import Path

import openpyxl

# Backups deliberately live off the OneDrive sync path — the workbook must
# never end up anywhere it could be published.
BACKUP_DIR = Path.home() / 'Library' / 'Application Support' / 'IMResidentDashboard' / 'workbook-backups'

KEEP_BACKUPS = 10


class WorkbookError(RuntimeError):
    """A guarded save was refused or rolled back. The workbook is unchanged."""


class WorkbookBusy(WorkbookError):
    """The workbook is open in Excel — writing now risks losing the save."""


class WorkbookConflict(WorkbookError):
    """The file changed on disk since we read it — refuse to clobber it."""


class WorkbookVerificationError(WorkbookError):
    """The workbook lost rows during the save; the backup was restored."""


def lock_file(path):
    """Path of the '~$' owner file Excel creates while a workbook is open."""
    return Path(path).parent / f'~${Path(path).name}'


def is_open_in_excel(path):
    return lock_file(path).exists()


def file_signature(path):
    """(mtime_ns, size) — captured before a read, rechecked before the write."""
    st = Path(path).stat()
    return (st.st_mtime_ns, st.st_size)


def backup_workbook(path, keep=KEEP_BACKUPS):
    """Copy the workbook into BACKUP_DIR, pruning all but the newest `keep`."""
    path = Path(path)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)

    stamp = datetime.now().strftime('%Y%m%d-%H%M%S')
    target = BACKUP_DIR / f'{path.stem}.{stamp}{path.suffix}'
    shutil.copy2(path, target)

    existing = sorted(BACKUP_DIR.glob(f'{path.stem}.*{path.suffix}'))
    for stale in existing[:-keep]:
        stale.unlink()

    return target


def sheet_row_counts(path):
    """{sheet name: count of non-empty rows} — used to prove nothing was lost.

    Read with data_only=False so formula cells count by their formula rather
    than their cached value. The cache is present right after Excel saves and
    absent right after openpyxl does, so counting cached values would make the
    700 XLOOKUP rows in OtherPoints appear and vanish on their own and trip the
    row-loss check on a perfectly good save.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        counts = {}
        for name in wb.sheetnames:
            counts[name] = sum(
                1 for row in wb[name].iter_rows(values_only=True)
                if any(cell is not None for cell in row)
            )
        return counts
    finally:
        wb.close()


def save_workbook(wb, path, expect_signature=None, allow_shrink=(), keep_backups=KEEP_BACKUPS):
    """Back up, atomically replace, then verify no sheet lost rows.

    expect_signature: the file_signature() taken before the workbook was read.
                      If the file has changed since, raise WorkbookConflict
                      rather than overwrite someone else's edit.
    allow_shrink:     sheet names that are legitimately rewritten and may end
                      up with fewer rows than before.
    """
    path = Path(path)

    if is_open_in_excel(path):
        raise WorkbookBusy(
            f'{path.name} is currently open in Excel ({lock_file(path).name} present). '
            'Close it and re-run — writing now risks the save being reverted.'
        )

    if expect_signature is not None and file_signature(path) != expect_signature:
        raise WorkbookConflict(
            f'{path.name} changed on disk while this script was running (another '
            'edit, or a OneDrive sync). Nothing was written — re-run to pick up '
            'the current version.'
        )

    before = sheet_row_counts(path)
    backup = backup_workbook(path, keep=keep_backups)

    # Write to a temp file on the same filesystem, then swap it in with a single
    # atomic rename, so the workbook is never left half-written.
    fd, tmp_name = tempfile.mkstemp(dir=str(path.parent), prefix=f'.{path.stem}.', suffix=path.suffix)
    os.close(fd)
    tmp = Path(tmp_name)
    try:
        wb.save(tmp)
        shutil.copystat(path, tmp)
        os.replace(tmp, path)
    except BaseException:
        tmp.unlink(missing_ok=True)
        raise

    after = sheet_row_counts(path)
    shrunk = {
        name: (before[name], after.get(name, 0))
        for name in before
        if name not in allow_shrink and after.get(name, 0) < before[name]
    }
    if shrunk:
        shutil.copy2(backup, path)
        detail = ', '.join(f'{n}: {b} -> {a}' for n, (b, a) in shrunk.items())
        raise WorkbookVerificationError(
            f'Save lost rows ({detail}). Restored the backup from {backup}.'
        )

    return backup
