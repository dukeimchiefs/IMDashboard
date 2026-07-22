#!/usr/bin/env python3
"""
Duke IM Resident Dashboard — Attendance Importer
=================================================
Downloads the email-free attendance export through Cloudflare Access and
appends any new records into the "Attendance" sheet of the shared
Point_Spreadsheet.xlsx workbook. Run refresh_data.py afterward to fold these
into the dashboard.

SETUP (one-time)
----------------
1. Install dependencies:
       pip3 install -r requirements.txt

2. Credentials are read from environment variables first. On the dashboard
   Mac, they fall back to the three macOS Keychain entries documented in the
   README. Never put real credentials in a committed file.

3. Run it whenever you want to sync attendance:
       python3 scrape_attendance.py

4. Run refresh_data.py afterward and commit the updated data.js.
"""

import os
import subprocess
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit('openpyxl not found. Install it with:  pip3 install -r requirements.txt')

try:
    from dotenv import load_dotenv
except ImportError:
    sys.exit('python-dotenv not found. Install it with:  pip3 install -r requirements.txt')

try:
    import requests
except ImportError:
    sys.exit('requests not found. Install it with:  pip3 install -r requirements.txt')

# ============================================================
# CONFIGURATION — edit these to match your setup
# ============================================================

ATTENDANCE_URL = 'https://imresidentdashboardapp.pages.dev/export'

# Same workbook refresh_data.py reads from.
EXCEL_FILE = Path('/Users/nbrazeau/Library/CloudStorage/OneDrive-SharedLibraries-DukeUniversity/Duke Chiefs 2024-2026 - Documents/zChief_Gamification/Point_Spreadsheet.xlsx')

ATTENDANCE_SHEET_NAME = 'AttendancePoints'

# ============================================================


KEYCHAIN_ACCOUNT = 'nbrazeau'
KEYCHAIN_SERVICES = {
    'ADMIN_EXPORT_KEY': 'imresidentdashboardapp-admin-export-key',
    'CF_ACCESS_CLIENT_ID': 'imresidentdashboard-access-client-id',
    'CF_ACCESS_CLIENT_SECRET': 'imresidentdashboard-access-client-secret',
}

EVENT_LABELS = {
    'noon_conference': 'Noon Conference',
    'learning_session': 'Learning Session',
    'grand_rounds': 'Grand Rounds',
    'welcome': 'Welcome',
}


def keychain_value(service):
    """Read a credential from macOS Keychain without echoing it."""
    if sys.platform != 'darwin':
        return None
    result = subprocess.run(
        ['security', 'find-generic-password', '-s', service, '-a', KEYCHAIN_ACCOUNT, '-w'],
        capture_output=True,
        text=True,
        check=False,
    )
    return result.stdout.strip() if result.returncode == 0 else None


def get_credentials():
    load_dotenv()
    credentials = {}
    missing = []
    for name, service in KEYCHAIN_SERVICES.items():
        value = os.environ.get(name) or keychain_value(service)
        if value:
            credentials[name] = value
        else:
            missing.append(name)
    if missing:
        sys.exit(
            '\nERROR: Missing attendance export credentials: '
            + ', '.join(missing)
            + '.\nSet them as environment variables or install the documented macOS Keychain entries.\n'
        )
    return credentials


def scrape_attendance(credentials):
    """Download the protected export and return (date, name, event) tuples."""
    try:
        response = requests.get(
            ATTENDANCE_URL,
            headers={
                'Accept': 'application/json',
                'X-Admin-Key': credentials['ADMIN_EXPORT_KEY'],
                'CF-Access-Client-Id': credentials['CF_ACCESS_CLIENT_ID'],
                'CF-Access-Client-Secret': credentials['CF_ACCESS_CLIENT_SECRET'],
            },
            timeout=30,
        )
    except requests.RequestException as error:
        raise RuntimeError('attendance export request failed') from error

    if response.status_code != 200:
        raise RuntimeError(f'attendance export returned HTTP {response.status_code}')
    try:
        payload = response.json()
    except requests.JSONDecodeError as error:
        raise ValueError('attendance export did not return JSON') from error

    if payload.get('ok') is not True or not isinstance(payload.get('rows'), list):
        raise ValueError('attendance export returned an unexpected response')

    rows = []
    for item in payload['rows']:
        if not isinstance(item, dict):
            raise ValueError('attendance export contained an invalid row')
        date = str(item.get('event_date', '')).strip()
        name = str(item.get('name', '')).strip()
        event_type = str(item.get('event_type', '')).strip()
        if not date or not name or event_type not in EVENT_LABELS:
            raise ValueError('attendance export contained a missing or unknown field')
        rows.append((date, name, EVENT_LABELS[event_type]))
    return rows


def append_to_workbook(path, sheet_name, scraped_rows):
    """Append new (date, name, event) rows to the Attendance sheet, skipping duplicates."""
    wb = openpyxl.load_workbook(path)

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        existing = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 3:
                existing.add((str(row[0]).strip(), str(row[1]).strip(), str(row[2]).strip()))
    else:
        ws = wb.create_sheet(sheet_name)
        ws.append(['Date', 'Name', 'Event'])
        existing = set()

    appended = 0
    skipped = 0
    for date, name, event in scraped_rows:
        key = (date, name, event)
        if key in existing:
            skipped += 1
            continue
        ws.append([date, name, event])
        existing.add(key)
        appended += 1

    wb.save(path)
    return appended, skipped


if __name__ == '__main__':
    if not EXCEL_FILE.exists():
        sys.exit(
            f'\nERROR: Excel file not found:\n  {EXCEL_FILE}\n\n'
            f'Edit EXCEL_FILE in scrape_attendance.py to point to your workbook.\n'
        )

    credentials = get_credentials()

    try:
        scraped_rows = scrape_attendance(credentials)
    except Exception as e:
        sys.exit(f'\nERROR downloading attendance export: {e}\n')

    if not scraped_rows:
        sys.exit('No attendance rows found on the page — nothing to append.\n')

    appended, skipped = append_to_workbook(EXCEL_FILE, ATTENDANCE_SHEET_NAME, scraped_rows)
    print(
        f'Scraped {len(scraped_rows)} row(s) — '
        f'appended {appended} new, skipped {skipped} already present.'
    )
